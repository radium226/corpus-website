from pathlib import Path
from openpyxl import load_workbook
import pendulum as p
import dateparser as dp
from PIL import Image

from ..models import *


def parse_date(text: str) -> p.Date | None:
    # return p.instance(dp.parse(text))
    try:
        return p.parse(text, format="%Y")
    except:
        return None

# - [x] Miniature
# - [x] Date
# - [x] Auteur
# - [x] Profession
# - [x] Pays
# - [x] Titre
# - [x] Source
# - [x] Medium
# - [x] Publication
# - [x] Thème(s) critique(s) (idées primaires)
# - [x] Paradigme esthétique (idées primaires)
# - [x] Motifs esthétiques
# - [x] Registre(s)
# - [x] Interprétation


class Store():

    def __init__(self, excel_file_path: Path):
        self.excel_file_path = excel_file_path
    
    def list_documents(self) -> list[Document]:
        workbook = load_workbook(filename=str(self.excel_file_path), rich_text=True)
        worksheet = workbook.active
        return [
            Document(
                thumbnail=Thumbnail(Image.open(Path(__file__).parent / "placeholder.png")),
                date=parse_date(str(row[1])),
                author=Author(row[2]),
                job=Job(row[3]),
                country=Country(row[4]),
                title=Title(row[5]),
                source=Source(row[6]),
                medium=Medium(row[7]),
                publication=Publication(row[8]),
                criticized_themes=[CriticizedTheme(row[9])],
                aesthetic_paradigms=[AestheticParadigm(row[10])],
                aesthetic_patterns=[AestheticPattern(row[11])],
                registers=[Register(row[12])],
                interpretation=Interpretation(row[13]),
            )
            for row in worksheet.iter_rows(min_row=2, values_only=True)
        ]
            