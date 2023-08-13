from openpyxl import load_workbook
from pprint import pprint
from pytest import mark


@mark.skip()
def test_load_xlsx():
    workbook = load_workbook(filename="./corpus/corpus.xlsx", rich_text=True)
    worksheet = workbook.active
    for row in worksheet.iter_rows():
        cell = row[0]
        print(cell.data_type)